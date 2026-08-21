import streamlit as st
import pandas as pd
import openpyxl
import os
import io
import re
import datetime

st.set_page_config(page_title="Form Suket Lahir", layout="wide")
st.title("📋 Form Input Data - SUKET LAHIR")

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "SUKET LAHIR.xlsx")
TODAY = datetime.date.today()
CURRENT_YEAR = TODAY.year

@st.cache_data
def load_excel_data(file_path):
    return pd.read_excel(file_path, sheet_name="ISIAN DATA", header=None)

def get_val(df, r, c):
    try:
        val = df.iloc[r, c]
        return "" if pd.isna(val) else str(val).strip()
    except:
        return ""

# --- FUNGSI UPDATE SELURUH DATA KE EXCEL ---
def update_excel_file(fd):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "ISIAN DATA" in wb.sheetnames:
        ws = wb["ISIAN DATA"]
        # Header & KK
        ws.cell(row=2, column=1, value=fd["no_surat"])
        ws.cell(row=6, column=5, value=fd["nama_kk"])
        ws.cell(row=7, column=5, value=fd["no_kk"])
        
        # Bayi
        ws.cell(row=11, column=5, value=fd["nama_bayi"])
        ws.cell(row=12, column=5, value=fd["jk_bayi"])
        ws.cell(row=14, column=5, value=fd["tempat_lahir_b"])
        ws.cell(row=16, column=5, value=fd["jam_lahir_b"])
        ws.cell(row=16, column=6, value=fd["tgl_b"])
        ws.cell(row=16, column=7, value=fd["bln_b"])
        ws.cell(row=16, column=8, value=fd["thn_b"])
        ws.cell(row=16, column=9, value=fd["umur_b"])
        ws.cell(row=18, column=5, value=fd["kelahiran_ke"])
        ws.cell(row=19, column=5, value=fd["penolong"])
        ws.cell(row=20, column=5, value=fd["berat_b"])
        ws.cell(row=21, column=5, value=fd["panjang_b"])

        # Ibu
        ws.cell(row=24, column=5, value=fd["nik_ibu"])
        ws.cell(row=25, column=5, value=fd["nama_ibu"])
        ws.cell(row=26, column=6, value=fd["tgl_i"])
        ws.cell(row=26, column=7, value=fd["bln_i"])
        ws.cell(row=26, column=8, value=fd["thn_i"])
        ws.cell(row=26, column=9, value=fd["umur_i"])
        ws.cell(row=27, column=5, value=fd["pekerjaan_i"])
        ws.cell(row=28, column=5, value=fd["alamat_i"])
        ws.cell(row=32, column=5, value=fd["kebangsaan_i"])
        ws.cell(row=34, column=6, value=fd["tgl_kawin"])
        ws.cell(row=34, column=7, value=fd["bln_kawin"])
        ws.cell(row=34, column=8, value=fd["thn_kawin"])

        # Ayah
        ws.cell(row=36, column=5, value=fd["nik_ayah"])
        ws.cell(row=37, column=5, value=fd["nama_ayah"])
        ws.cell(row=39, column=6, value=fd["tgl_a"])
        ws.cell(row=39, column=7, value=fd["bln_a"])
        ws.cell(row=39, column=8, value=fd["thn_a"])
        ws.cell(row=39, column=9, value=fd["umur_a"])
        ws.cell(row=39, column=5, value=fd["pekerjaan_a"])
        ws.cell(row=40, column=5, value=fd["alamat_a"])
        ws.cell(row=44, column=5, value=fd["kebangsaan_a"])

        # Pelapor
        ws.cell(row=47, column=5, value=fd["nik_pelapor"])
        ws.cell(row=48, column=5, value=fd["nama_pelapor"])
        ws.cell(row=49, column=5, value=fd["umur_pelapor"])
        ws.cell(row=50, column=5, value=fd["jk_pelapor"])
        ws.cell(row=51, column=5, value=fd["pekerjaan_pelapor"])
        ws.cell(row=52, column=5, value=fd["alamat_pelapor"])

        # Saksi I & II
        ws.cell(row=57, column=5, value=fd["nik_s1"])
        ws.cell(row=58, column=5, value=fd["nama_s1"])
        ws.cell(row=59, column=5, value=fd["umur_s1"])
        ws.cell(row=61, column=5, value=fd["alamat_s1"])

        ws.cell(row=66, column=5, value=fd["nik_s2"])
        ws.cell(row=67, column=5, value=fd["nama_s2"])
        ws.cell(row=68, column=5, value=fd["umur_s2"])
        ws.cell(row=70, column=5, value=fd["alamat_s2"])

        # Keterangan
        ws.cell(row=72, column=5, value=fd["tgl_surat"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if not os.path.exists(EXCEL_FILE):
    st.error(f"⚠️ File Excel tidak ditemukan: {EXCEL_FILE}")
else:
    try:
        df = load_excel_data(EXCEL_FILE)

        st.subheader("📝 Form Isian Data SUKET LAHIR")
        with st.form("form_suket_lengkap"):
            st.markdown("##### 1. Header & Data Kepala Keluarga")
            c1, c2, c3 = st.columns(3)
            with c1: no_surat = st.text_input("Nomor Surat", value=get_val(df, 1, 0))
            with c2: nama_kk = st.text_input("Nama Kepala Keluarga", value=get_val(df, 5, 4))
            with c3: no_kk = st.text_input("Nomor Kartu Keluarga (KK)", value=get_val(df, 6, 4))

            st.divider()

            st.markdown("##### 2. Data Bayi / Anak")
            c1, c2, c3 = st.columns(3)
            with c1:
                nama_bayi = st.text_input("Nama Bayi", value=get_val(df, 10, 4))
                jk_bayi = st.text_input("Jenis Kelamin Bayi", value=get_val(df, 11, 4))
                tempat_lahir_b = st.text_input("Tempat Kelahiran", value=get_val(df, 13, 4))
            with c2:
                jam_lahir_b = st.text_input("Jam Lahir", value=get_val(df, 15, 4))
                st.caption("Tanggal Lahir Bayi (TGL / BLN / THN / UMUR)")
                cb1, cb2, cb3, cb4 = st.columns(4)
                with cb1: tgl_b = st.text_input("Tgl", value=get_val(df, 15, 5), key="b_tgl")
                with cb2: bln_b = st.text_input("Bln", value=get_val(df, 15, 6), key="b_bln")
                with cb3: thn_b = st.text_input("Thn", value=get_val(df, 15, 7), key="b_thn")
                
                # Hitung Umur Bayi Otomatis
                default_umur_b = get_val(df, 15, 8)
                calc_b = default_umur_b
                if thn_b.isdigit():
                    try:
                        d_b = int(tgl_b) if tgl_b.isdigit() else 1
                        m_b = int(bln_b) if bln_b.isdigit() else 1
                        y_b = int(thn_b)
                        tgl_lahir_bayi = datetime.date(y_b, m_b, d_b)
                        selisih_hari = (TODAY - tgl_lahir_bayi).days

                        if selisih_hari >= 0:
                            if selisih_hari < 30:
                                calc_b = f"{selisih_hari} Hari"
                            elif selisih_hari < 365:
                                calc_b = f"{selisih_hari // 30} Bulan"
                            else:
                                calc_b = f"{CURRENT_YEAR - y_b} Tahun"
                        else:
                            calc_b = "0 Hari"
                    except:
                        calc_b = str(CURRENT_YEAR - int(thn_b)) if thn_b.isdigit() else default_umur_b

                with cb4: umur_b = st.text_input("Umur", value=calc_b, key="b_umur")
                kelahiran_ke = st.text_input("Kelahiran Ke-", value=get_val(df, 17, 4))
            with c3:
                penolong = st.text_input("Penolong Kelahiran", value=get_val(df, 18, 4))
                berat_b = st.text_input("Berat Bayi (Gram)", value=get_val(df, 19, 4))
                panjang_b = st.text_input("Panjang Bayi (Cm)", value=get_val(df, 20, 4))

            st.divider()

            col_ibu, col_ayah = st.columns(2)
            with col_ibu:
                st.markdown("##### 3. Data Ibu")
                nik_ibu = st.text_input("NIK Ibu", value=get_val(df, 23, 4))
                nama_ibu = st.text_input("Nama Lengkap Ibu", value=get_val(df, 24, 4))
                st.caption("Tanggal Lahir Ibu (TGL / BLN / THN / UMUR)")
                ci1, ci2, ci3, ci4 = st.columns(4)
                with ci1: tgl_i = st.text_input("Tgl", value=get_val(df, 25, 5), key="i_tgl")
                with ci2: bln_i = st.text_input("Bln", value=get_val(df, 25, 6), key="i_bln")
                with ci3: thn_i = st.text_input("Thn", value=get_val(df, 25, 7), key="i_thn")
                
                # Hitung Umur Ibu Otomatis jika Thn diisi
                default_umur_i = get_val(df, 25, 8)
                if thn_i.isdigit():
                    calc_i = str(CURRENT_YEAR - int(thn_i))
                else:
                    calc_i = default_umur_i

                with ci4: umur_i = st.text_input("Umur", value=calc_i, key="i_umur")
                
                pekerjaan_i = st.text_input("Pekerjaan Ibu", value=get_val(df, 26, 4))
                alamat_i = st.text_input("Alamat Ibu", value=get_val(df, 27, 4))
                kebangsaan_i = st.text_input("Kebangsaan Ibu", value=get_val(df, 31, 4))
                st.caption("Tanggal Perkawinan (TGL / BLN / THN)")
                cp1, cp2, cp3 = st.columns(3)
                with cp1: tgl_kawin = st.text_input("Tgl Kawin", value=get_val(df, 33, 5))
                with cp2: bln_kawin = st.text_input("Bln Kawin", value=get_val(df, 33, 6))
                with cp3: thn_kawin = st.text_input("Thn Kawin", value=get_val(df, 33, 7))

            with col_ayah:
                st.markdown("##### 4. Data Ayah")
                nik_ayah = st.text_input("NIK Ayah", value=get_val(df, 35, 4))
                nama_ayah = st.text_input("Nama Lengkap Ayah", value=get_val(df, 36, 4))
                st.caption("Tanggal Lahir Ayah (TGL / BLN / THN / UMUR)")
                ca1, ca2, ca3, ca4 = st.columns(4)
                with ca1: tgl_a = st.text_input("Tgl", value=get_val(df, 38, 5), key="a_tgl")
                with ca2: bln_a = st.text_input("Bln", value=get_val(df, 38, 6), key="a_bln")
                with ca3: thn_a = st.text_input("Thn", value=get_val(df, 38, 7), key="a_thn")
                
                # Hitung Umur Ayah Otomatis jika Thn diisi
                default_umur_a = get_val(df, 38, 8)
                if thn_a.isdigit():
                    calc_a = str(CURRENT_YEAR - int(thn_a))
                else:
                    calc_a = default_umur_a

                with ca4: umur_a = st.text_input("Umur", value=calc_a, key="a_umur")

                pekerjaan_a = st.text_input("Pekerjaan Ayah", value=get_val(df, 38, 4))
                alamat_a = st.text_input("Alamat Ayah", value=get_val(df, 39, 4))
                kebangsaan_a = st.text_input("Kebangsaan Ayah", value=get_val(df, 43, 4))

            st.divider()

            st.markdown("##### 5. Data Pelapor")
            c1, c2, c3 = st.columns(3)
            with c1:
                nik_pelapor = st.text_input("NIK Pelapor", value=get_val(df, 46, 4))
                nama_pelapor = st.text_input("Nama Lengkap Pelapor", value=get_val(df, 47, 4))
            with c2:
                umur_pelapor = st.text_input("Umur Pelapor", value=get_val(df, 48, 4))
                jk_pelapor = st.text_input("Jenis Kelamin Pelapor", value=get_val(df, 49, 4))
            with c3:
                pekerjaan_pelapor = st.text_input("Pekerjaan Pelapor", value=get_val(df, 50, 4))
                alamat_pelapor = st.text_input("Alamat Pelapor", value=get_val(df, 51, 4))

            st.divider()

            col_s1, col_s2 = st.columns(2)
            with col_s1:
                st.markdown("##### 6. Data Saksi I")
                nik_s1 = st.text_input("NIK Saksi I", value=get_val(df, 56, 4))
                nama_s1 = st.text_input("Nama Lengkap Saksi I", value=get_val(df, 57, 4))
                umur_s1 = st.text_input("Umur Saksi I", value=get_val(df, 58, 4))
                alamat_s1 = st.text_input("Alamat Saksi I", value=get_val(df, 60, 4))

            with col_s2:
                st.markdown("##### 7. Data Saksi II")
                nik_s2 = st.text_input("NIK Saksi II", value=get_val(df, 65, 4))
                nama_s2 = st.text_input("Nama Lengkap Saksi II", value=get_val(df, 66, 4))
                umur_s2 = st.text_input("Umur Saksi II", value=get_val(df, 67, 4))
                alamat_s2 = st.text_input("Alamat Saksi II", value=get_val(df, 69, 4))

            st.divider()

            st.markdown("##### 8. Keterangan Surat")
            tgl_surat = st.text_input("Tempat & Tanggal Surat", value=get_val(df, 71, 4))

            btn = st.form_submit_button("🔄 Simpan Perubahan Form")

        # Dictionary penampung data
        fd = {
            "no_surat": no_surat, "nama_kk": nama_kk, "no_kk": no_kk,
            "nama_bayi": nama_bayi, "jk_bayi": jk_bayi, "tempat_lahir_b": tempat_lahir_b,
            "jam_lahir_b": jam_lahir_b, "tgl_b": tgl_b, "bln_b": bln_b, "thn_b": thn_b, "umur_b": umur_b,
            "kelahiran_ke": kelahiran_ke, "penolong": penolong, "berat_b": berat_b, "panjang_b": panjang_b,
            "nik_ibu": nik_ibu, "nama_ibu": nama_ibu, "tgl_i": tgl_i, "bln_i": bln_i, "thn_i": thn_i, "umur_i": umur_i,
            "pekerjaan_i": pekerjaan_i, "alamat_i": alamat_i, "kebangsaan_i": kebangsaan_i,
            "tgl_kawin": tgl_kawin, "bln_kawin": bln_kawin, "thn_kawin": thn_kawin,
            "nik_ayah": nik_ayah, "nama_ayah": nama_ayah, "tgl_a": tgl_a, "bln_a": bln_a, "thn_a": thn_a, "umur_a": umur_a,
            "pekerjaan_a": pekerjaan_a, "alamat_a": alamat_a, "kebangsaan_a": kebangsaan_a,
            "nik_pelapor": nik_pelapor, "nama_pelapor": nama_pelapor, "umur_pelapor": umur_pelapor,
            "jk_pelapor": jk_pelapor, "pekerjaan_pelapor": pekerjaan_pelapor, "alamat_pelapor": alamat_pelapor,
            "nik_s1": nik_s1, "nama_s1": nama_s1, "umur_s1": umur_s1, "alamat_s1": alamat_s1,
            "nik_s2": nik_s2, "nama_s2": nama_s2, "umur_s2": umur_s2, "alamat_s2": alamat_s2,
            "tgl_surat": tgl_surat
        }

        nama_bayi_clean = re.sub(r'[^a-zA-Z0-9_-]', '_', nama_bayi).strip('_') if nama_bayi else "BAYI"
        nama_file_excel = f"SUKET_LAHIR_{nama_bayi_clean}.xlsx"

        st.divider()
        st.subheader("📥 Unduh Berkas Excel Ter-update")

        st.download_button(
            label=f"📥 Download File Excel Utuh ({nama_file_excel})",
            data=update_excel_file(fd),
            file_name=nama_file_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    except Exception as e:
        st.error(f"Error: {e}")
